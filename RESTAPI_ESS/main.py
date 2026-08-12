
# ============================================================================
# main.py - Oracle Fusion ESS Job Automation via REST API
# ============================================================================
"""
WHAT THIS FILE DOES:
    Automates submission and monitoring of Oracle Fusion ESS (Enterprise Scheduling Service)
    jobs using the REST API. Reads job definitions from Excel files and submits them to
    Oracle Fusion with proper scheduling, parameters, and credentials.

KEY FEATURES:
    - Multi-tenant support (different credentials per row)
    - On-demand and scheduled job submission
    - Duplicate detection (won't submit if already running)
    - Poll-to-completion for on-demand jobs
    - Excel audit logging (tracks all submissions)
    - Text file logging (detailed console + file logs)
    - DRY_RUN mode for testing

USAGE:
    # Default mode - process data.xlsx from drop folder:
    python main.py
    
    # Scenario mode - process predefined file list:
    python main.py --scenario P2T
    
    # Folder mode - process all Excel files in scenarios/monthly/:
    python main.py --folder monthly
    
    # Folder mode with strict alphabetical order:
    python main.py --folder monthly --strict-order

    #FOR OUR CASE SPECIFIC SCENARIO HAS BEEN CREATED

EXCEL FILE STRUCTURE:
    Required columns:
    - Display Name: Human-readable job name OR use jobDefinitionName
    - application: Oracle module (e.g., "ERP", "HCM", "SCM")
    - login: Oracle username (or "scheduler" for password file)
    - password: Oracle password (optional if using password file)
    - jobDefinitionId: Full job path (e.g., "JobDefinition://oracle/apps/ess/custom/MyJob")
    - jobPackageName: Package path (e.g., "/oracle/apps/ess/custom")
    - jobDefinitionName: Short job name (e.g., "MyJob")
    - description: Job description for audit trail
    - schedule_type: "on_demand" or "recurring"
    - startDate, endDate, recurrencePattern: For scheduled jobs
    - argument1 through argument30: Job parameters
    - Any named parameter columns matching job prompts

DEPENDENCIES:
    - openpyxl: Excel file reading/writing
    - requests: HTTP REST API calls
    - python-dotenv: Environment variable loading
    - Custom modules:
        - logger.py: Console + file logging
        - fusion_api.py: ESSClient wrapper for Oracle REST API
        - job_audit_xlsx.py: Excel audit trail

ENVIRONMENT VARIABLES (.env):
    FUSION_BASEURL: Oracle Fusion base URL (required)
    DATA_XLSX_PATH: Path to main Excel file (optional, default: data.xlsx)
    DATA_DROP_DIR: Folder to watch for new Excel files (optional, default: project root)
    DATA_PREFER_REGEX: Regex to prefer certain filenames (optional)
    DATA_STABLE_SEC: Seconds to wait for file stability (optional, default: 2)
    SCENARIOS_ROOT: Root folder for scenario subfolders (optional, default: scenarios/)

PASSWORD FILE:
    Alternative to storing passwords in Excel:
    Create ./password/password.txt with scheduler password
    Use "scheduler" in Excel login column to trigger password file lookup

KNOWN ISSUES:
    - OSCS jobs (Search/Knowledge) require special parameter handling
      (fan-out to multiple parameter aliases)
    - Some Oracle pods block catalog access (job resolution by display name fails)
    - Scheduled jobs return parent request ID (child executions tracked separately)

DEVELOPMENT NOTES:
    Last Updated: February 2026
    Developer: Alex Cacaras
    
    Originally built for single-file mode, later expanded to support
    scenarios and folder modes for more flexible batch processing.

 ============================================================================
 UPDATE 1.0.1 - SMART RETRY SYSTEM (March 2026)
 ============================================================================

CHANGES IN VERSION 1.0.1:
    - Added timeout-aware polling for on-demand jobs
    - Implemented job status tracking via JSON file
    - Added configurable timeout thresholds (overall + WAIT state)
    - Display name tracking for UI retry matching
    - ACL job failure detection and alerts
    
NEW CONFIGURATION:
    MAX_POLL_TIME_SECONDS: Maximum time to wait for job completion (default: 600s)
    POLL_INTERVAL_SECONDS: How often to check job status (default: 10s)
    MAX_WAIT_STATE_TIME: Maximum time job can stay in WAIT state (default: 300s)
    STATUS_FILE: Path to job_status_tracker.json for UI retry coordination

HOW IT WORKS:
    1. REST API submits jobs and monitors with timeout thresholds
    2. Failed/timed-out jobs are logged to job_status_tracker.json
    3. System returns code 1 if any jobs failed (triggers UI retry in Ui_Automation.py)
    4. UI automation reads JSON and retries only failed jobs via Oracle UI
    5. ACL jobs (critical security jobs) are tracked separately with alerts

 ============================================================================
 END UPDATE 1.0.1
 ============================================================================
 ============================================================================
 UPDATE 1.0.3 - (client) POD FIXES + LOGGING ACCURACY (March 2026)
 ============================================================================
 
CHANGES IN VERSION 1.0.3:
 
    DUPLICATE DETECTION:
    - On-demand jobs now pass argument1 to find_active_request_for_definition()
      so OSCS jobs with different index names are never falsely blocked
    - Scheduled jobs use new find_active_scheduled_request() matching startDate
      and argument1 for accurate duplicate detection
 
    LOGGING / STATUS TRACKER:
    - total_rows now increments at top of loop — counts every attempted row
      (previously only counted rows that reached submit, causing total_jobs
      to undercount while failed overcounted — e.g. failed:25 > total_jobs:17)
    - EXISTING_RUNNING / PAUSED / WAIT after short poll window → skipped not
      failed (jobs in flight on Oracle are not failures) At end the code will check the final status
      many of these skipped jobs will come up as succeeded.
    - All add_failed_job() calls use display_name so job_status_tracker.json
      names match Excel Display Name for ui_ess_jobs.py retry matching
    - ACL_JOB_NAMES now includes jobDefinitionNames (ComputeUsersACLProcessor,
      DataSecurityAclRefresh, ManageExcludedUsersACL) so ACL detection works
      regardless of which name format is in the JSON
 
    POLLING:
    - WAIT / PAUSED / BLOCKED / RETRYING treated as normal Oracle queuing states
      (Postman confirmed PAUSED → SUCCEEDED naturally)
    - Removed MAX_WAIT_STATE_TIME early-exit — only overall 600s timeout applies
 
    SUBMISSION:
    - jobDefinitionId tried first on every submit (fixes ESS-11003 error)
 
 ============================================================================
 END UPDATE 1.0.3
 ============================================================================
"""
from __future__ import annotations  # must be the very first import, for simplifying type hinting throughout file
# Inside RESTAPI_ESS/main.py
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))
import os, json, time, re, shutil, argparse, sys  #os for os features like reading env variables or file locations, json for parse/create schedule jsons, time for time related
# re for regular expressions, shutil for copying Excel files into main, argparse for simplifying reading command lines

from typing import Dict, List, Union  #typing for type hints and reading improvability
from pathlib import Path     #pathlib/path for handling file/directory paths

from dotenv import load_dotenv   # dotenv for loading env variables from .env file into python
from openpyxl import load_workbook  #openpyx1 for reading and writing Excel files

from logger import get_logger  # custom console+file logger
from fusion_api import ESSClient, ESSDuplicatePendingError   #custom that contains ESSClient class for communiction with Oracle Rest API
from job_audit_xlsx import log_job_run_xlsx  # for Excel logging of job runs
from datetime import datetime, timezone  # For timestamps
# setup/configuration
logger = get_logger(__name__)  # one logger for this module
load_dotenv(Path(__file__).resolve().parents[1] / ".env")

# Auto-wire: derive all env aliases from the single credentials
os.environ.setdefault("FUSION_BASEURL", os.getenv("TENANT_BASE_URL", ""))
for alias in (os.getenv("FUSION_USERS") or "").split(","):
    alias = alias.strip()
    if alias:
        os.environ.setdefault(f"FUSION_{alias}_LOGIN", os.getenv("FUSION_USERNAME", ""))
        os.environ.setdefault(f"FUSION_{alias}_PASSWORD", os.getenv("FUSION_PASSWORD", ""))
# Stable paths
PROJECT_ROOT = Path(__file__).resolve().parent
DATA_XLSX_PATH = Path(os.getenv("DATA_XLSX_PATH") or (PROJECT_ROOT / "data.XLSX"))

# Where Client drops their Excel file if name not data.XLSX it will be converted to that path
DROP_DIR = Path(os.getenv("DATA_DROP_DIR") or PROJECT_ROOT)

# prefer names that match this regex
PREFER_NAME_REGEX = os.getenv("DATA_PREFER_REGEX") or ""

# How long to wait for file to upload
try:
    DATA_STABLE_SEC = int(os.getenv("DATA_STABLE_SEC") or "2")
except ValueError:
    DATA_STABLE_SEC = 2
# ==== Polling timeout configuration can be changed to be faster ====
MAX_POLL_TIME_SECONDS =100   # 100 seconds max per job 
POLL_INTERVAL_SECONDS = 10     # Check status every 10 seconds
#MAX_WAIT_STATE_TIME = 5      # 5 minutes max stuck in WAIT

# ==== Status tracker file ====
STATUS_FILE = Path(__file__).parent / "job_status_tracker.json"

# ==== ACL job identifiers (CRITICAL jobs) ====
ACL_JOB_KEYWORDS = ["ACL", "Access Control", "Security", "Role Assignment"]
# ==== Multi-Excel support: scenario / folder modes ====

# Root folder that holds scenario subfolders (default: <project>/scenarios)
SCENARIOS_ROOT = Path(os.getenv("SCENARIOS_ROOT") or (PROJECT_ROOT / "scenarios"))

# Optional “official” scenarios (only used if you call --scenario) 
SCENARIO_FILES = {
    "P2T": ["(client)P2T.xlsx"], # change this for different Excel
    #"P2T": ["TESTING.xlsx"], #for testing
}

# --- Password-file override for scheduler (simple + backward-compatible) ---
PASSWORD_FILE = PROJECT_ROOT / "password" / "password.txt"   # one line: latest scheduler password

def _read_password_file() -> str | None:
    """Return password from ./password/password.txt if present & non-empty; else None.
    Hit or miss, I think it works fine with just env too"""
    try:
        if PASSWORD_FILE.exists():
            val = PASSWORD_FILE.read_text(encoding="utf-8").strip()
            return val or None
    except Exception:
        # never raise for password file; silently ignore and fall back to .env
        #This ensures the script doesn't crash if password file is missing
        pass
    return None


def _parse_args():
    """
    Parse command line arguments for different processing modes.
    
    MODES:
        --scenario: Run predefined file list (e.g., --scenario P2T)
        --folder: Process all Excel files in scenarios/<folder>/
        --strict-order: Folder mode only - alphabetical order instead of newest first
    """
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--scenario", choices=SCENARIO_FILES.keys(),
                    help="Run a predefined file list (strict names) like P2T or QuarterlyPatch")
    ap.add_argument("--folder", help="Process all Excel files under scenarios/<folder> (any names)")
    ap.add_argument("--strict-order", action="store_true",
                    help="Folder mode only: process files in alphabetical order (default: newest first)")
    ap.add_argument("--help", action="help")
    return ap.parse_args()

def _existing_file_in_drop(name: str) -> Path | None:
    #searches for Excel in multiple locations
    #file shouldn't start with $ because it is in use or not saved
    # Look in DROP_DIR first, then project root, then scenarios root
    for base in [DROP_DIR, PROJECT_ROOT, SCENARIOS_ROOT]:
        p = Path(base) / name
        if p.exists() and p.is_file() and not p.name.startswith("~$"):
            return p
    return None

def _collect_excel_paths(args) -> list[Path]:
    """Return a list of Excel file paths to process based on args.
       If no args provided, falls back to your current single-file behavior."""
    # 1) scenario mode — exact filenames (controlled runs)
    if args.scenario:
        out = []
        for fname in SCENARIO_FILES[args.scenario]:
            p = _existing_file_in_drop(fname)
            if p:
                out.append(p)
            else:
                logger.warning(f"[scenario:{args.scenario}] Missing file: {fname}")
        return out

    # 2) folder mode — any *.xls* under scenarios/<folder> (flexible runs)
    if args.folder:
        folder = SCENARIOS_ROOT / args.folder
        if not folder.exists():
            logger.error(f"Scenario folder not found: {folder}")
            return []
        cands = [p for p in folder.glob("*.xls*") if p.is_file() and not p.name.startswith("~$")]
        if not cands:
            logger.warning(f"No Excel files under {folder}")
            return []
        if args.strict_order:
            cands.sort(key=lambda p: p.name.lower())                  # predictable alpha
        else:
            cands.sort(key=lambda p: p.stat().st_mtime, reverse=True) # newest first
        return cands

    # 3) fallback — your existing single-drop behavior
    eff = refresh_data_xlsx_from_drop()
    return [eff] if eff else []


DRY_RUN = False  # True = print payloads, no network calls. False = real calls.

# read Oracle tenant & creds from .env. rstrip("/") avoids double slashes when we build URLs
BASEURL = (os.getenv("FUSION_BASEURL") or "").rstrip("/")
# USERNAME/PASSWORD are now chosen per Excel row; we only need BASEURL to be present.
if not BASEURL:
    raise SystemExit("Missing FUSION_BASEURL in .env")

# -------- Excel helpers --------
def _is_file_stable(path: Path, seconds: int) -> bool:
    """
    Check if file size has stabilized (not actively being written).
    WHY THIS EXISTS:
        When monitoring a drop folder, we need to know when file upload
        is complete before trying to process it. Reading a partially-uploaded
        file causes Excel corruption errors.
    """
    try:
        s1 = path.stat().st_size
        time.sleep(seconds)
        s2 = path.stat().st_size
        return s1 == s2
    except FileNotFoundError:
        return False

def _pick_latest_excel(drop_dir: Path, prefer_regex: str = "") -> Path | None:
    """
    Find the most recently modified Excel file in drop folder.
    
    FILTERS:
        - Only .xlsx or .xlsm files
        - Excludes Excel temp files (~$filename.xlsx)
        - Optionally prefers files matching regex pattern
    
    DESIGN DECISION - Prefer Regex:
        If DATA_PREFER_REGEX is set (e.g., ".*P2T.*"), files matching
        that pattern are considered first. This lets you prioritize
        specific filenames when multiple Excel files are present.
        Because code can run multiple Excels
    """
    cands = [
        p for p in drop_dir.glob("*")
        if p.is_file()
        and p.suffix.lower() in (".xlsx", ".xlsm")
        and not p.name.startswith("~$")
    ]
    if not cands:
        return None

    if prefer_regex:
        try:
            pat = re.compile(prefer_regex, re.IGNORECASE)
            preferred = [p for p in cands if pat.search(p.name)]
            if preferred:
                cands = preferred
        except re.error as e:
            logger.warning(f"Ignoring bad DATA_PREFER_REGEX: {e}")

    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return cands[0]

def refresh_data_xlsx_from_drop() -> Path | None:
    """
    Update data.xlsx from drop folder if new file is available.
    WHAT IT DOES:
        1. Find latest Excel file in DROP_DIR
        2. Check if it's stable (finished uploading)
        3. Copy to data.xlsx if newer than current
        4. Return path to data.xlsx (or None if unavailable)
    
    DESIGN DECISION - File Stability Check:
        Waits DATA_STABLE_SEC seconds to confirm file size isn't changing.
        Prevents processing partially-uploaded files.
    
    DESIGN DECISION - Preserve Existing:
        If no new file found, returns path to existing data.xlsx
        This allows script to run even when drop folder is empty.
    """
    try:
        latest = _pick_latest_excel(DROP_DIR, PREFER_NAME_REGEX)
        if not latest:
            if DATA_XLSX_PATH.exists():
                logger.info(f"No new Excel found; using existing {DATA_XLSX_PATH.name}.")
                return DATA_XLSX_PATH
            logger.warning(f"No Excel found in {DROP_DIR} and {DATA_XLSX_PATH} doesn’t exist.")
            return None

        if latest.resolve() == DATA_XLSX_PATH.resolve():
            if _is_file_stable(DATA_XLSX_PATH, DATA_STABLE_SEC):
                logger.info(f"Using existing {DATA_XLSX_PATH.name} (already latest).")
                return DATA_XLSX_PATH
            logger.warning(f"{DATA_XLSX_PATH.name} not yet stable; try again shortly.")
            return None

        if not _is_file_stable(latest, DATA_STABLE_SEC):
            logger.warning(f"Latest Excel not stable yet: {latest.name} (size still changing).")
            return None

        shutil.copy2(latest, DATA_XLSX_PATH)
        logger.info(f"Updated {DATA_XLSX_PATH.name} from drop file: {latest.name}")
        return DATA_XLSX_PATH
    except Exception as e:
        logger.exception(f"refresh_data_xlsx_from_drop failed: {e}")
        return None

def read_jobs_from_excel(path: Union[str, Path] = DATA_XLSX_PATH) -> List[Dict[str, str]]:
     
    #Read Excel file and convert to list of row dictionaries.
    
   # WHAT IT DOES:
   #     1. Load workbook with data_only=True (evaluates formulas)
   #     2. Read header row to get column names
   #     3. Convert each data row to dictionary (column name → value)
   #     4. Skip completely empty rows
    
  #  RETURNS:
 #       List of dictionaries, one per data row
 #       Empty list if file not found or no rows
    
#    DESIGN DECISION - read_only=True:
 #       Opens workbook in read-only mode for better performance
#       I am not modifying the file, just reading job definitions
    
  #  DESIGN DECISION - data_only=True:
 #       Evaluates formulas to their calculated values
 #       Without this, I'd get formula strings instead of results
 #       Example: "=A1+B1" becomes "150" (the actual calculated value)
    
 #   DESIGN DECISION - Skip Empty Rows:
 #       Checks if ALL cells in row are empty before skipping
 #       Allows for intentional blank rows in Excel for organization
    
    path = Path(path)
    if not path.exists():
        logger.error(f"Excel not found at: {path}")
        return []

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_idx = {i: h for i, h in enumerate(headers) if h}

    job_rows: List[Dict[str, str]] = []
    for r in rows[1:]:
        row_dict: Dict[str, str] = {}
        is_empty = True
        for i, cell in enumerate(r):
            if i in header_idx:
                val = "" if cell is None else str(cell).strip()
                row_dict[header_idx[i]] = val
                if val:
                    is_empty = False
        if not is_empty:
            job_rows.append(row_dict)

    return job_rows

# ---------------- Param extraction (robust) ----------------
RESERVED_COLS = {
    "Enabled","Display Name","jobDefinitionId","jobPackageName","jobDefinitionName","application",
    "startDate","endDate","icalString","schedule_json","className", "Username", "AllowDuplicate"
}
ARG_PREFIX = "argument"

_NULL_SENTINELS = {"null", "none", "n/a", "na"}  # treated as empty

def _nullish(v: str | None) -> bool:
    """
    Check if value should be treated as empty/null.
    
    WHAT IT DOES:
        Returns True if value is None, empty string, or null sentinel
        Returns False otherwise
    
    NULL SENTINELS:
        - "null" (case insensitive)
        - "none" (case insensitive)
        - "n/a" (case insensitive)
        - "na" (case insensitive)
    
    WHY THIS EXISTS:
        Excel users sometimes enter "N/A" or "null" instead of leaving cells empty
        We need to treat these as missing values, not literal strings
    """
    if v is None:
        return True
    s = str(v).strip()
    return (s == "") or (s.lower() in _NULL_SENTINELS)

# normalize headers like "Argument 2", "ARGUMENT2", "argument  2" → "argument2"
_ARG_HEADER_RX = re.compile(r"^argument\s*(\d{1,2})$", re.IGNORECASE)
def _normalize_arg_header(h: str) -> str | None:
    """
    Normalize argument header to standard format.
    
    WHAT IT DOES:
        Converts various argument column formats to lowercase "argumentN"
    
    EXAMPLES:
        "Argument 2" → "argument2"
        "ARGUMENT2" → "argument2"
        "argument  2" → "argument2"
        "ARG 5" → None (not a match)
    
    WHY THIS EXISTS:
        Excel users format argument columns inconsistently
        This ensures the code recognize them all as the same parameter
    """
    m = _ARG_HEADER_RX.match(h.strip())
    return f"argument{m.group(1)}" if m else None

def extract_params(row: Dict[str, str]) -> Dict[str, str]:
    """
    Robust parameter collection:
     WHAT IT DOES:
        1. Extract named prompts (exact case preserved)
        2. Extract argument1..30 (normalized, case-insensitive)
        3. Skip reserved columns (Display Name, application, etc.)
        4. Skip null/empty values

      DESIGN DECISION - Exact Case for Named Prompts:
        Oracle jobs expect exact parameter names
        "Business Unit" ≠ "business unit" ≠ "BUSINESS UNIT"
        We preserve whatever case user entered in Excel header
        I used to use named argument sin Excel not anymore just need arg1,2,3,etc but left in
    
    DESIGN DECISION - Case-Insensitive for argumentN:
        "Argument 2", "ARGUMENT2", "argument2" all map to "argument2"
        Makes Excel more user-friendly while maintaining compatibility
    
    DESIGN DECISION - Skip Nullish Values:
        Empty cells, "N/A", "null", etc. are not sent as parameters
        Prevents errors from Oracle expecting specific data types
    """
    params: Dict[str, str] = {}

    # 1) Named prompts
    for k, v in row.items():
        if _nullish(v):
            continue
        if not k:
            continue
        if k in RESERVED_COLS:
            continue
        if _normalize_arg_header(k):
            continue
        params[k.strip()] = str(v).strip()

    # 2) Generic arguments (argument1..30)
    for n in range(1, 31):
        key = f"argument{n}"
        raw = row.get(key, "")
        if _nullish(raw):
            for hdr, val in row.items():
                if _normalize_arg_header(hdr) == key:
                    raw = val
                    break
        if _nullish(raw):
            continue
        params[key] = str(raw).strip()

    return params

def extract_schedule(row: Dict[str, str]) -> Dict:
    """
    Extract schedule information from Excel row.
    
    TWO FORMATS SUPPORTED:
        1. schedule_json: Full JSON schedule object (advanced users)
        2. icalString + startDate/endDate: Individual columns (most users)
    
    PRIORITY:
        schedule_json takes precedence if present
        Falls back to icalString format if schedule_json empty
    
    EXAMPLES:
        # Format 1 (JSON):
        schedule_json: {"recurrences": [{"icalString": "FREQ=DAILY", "startDate": "2025-01-01"}]}
        
        # Format 2 (Individual columns):
        icalString: FREQ=DAILY
        startDate: 2025-01-01T00:00:00Z
        endDate: 2025-12-31T23:59:59Z
    
    RETURNS:
        Empty dict {} for on-demand (immediate) jobs
        Nested dict with recurrences array for scheduled jobs
    
    DESIGN DECISION - JSON Format First:
        Advanced users can copy/paste exact JSON from Oracle UI
        Provides maximum flexibility for complex schedules
    
    DESIGN DECISION - Individual Columns Fallback:
        Most users prefer simple columns over JSON
        Easier to understand and modify in Excel
    """
    sch_json = (row.get("schedule_json") or "").strip()
    if sch_json:
        obj = json.loads(sch_json)
        if not isinstance(obj, dict):
            raise ValueError("schedule_json must be a JSON object")
        return obj

    ical = (row.get("icalString") or "").strip()
    start_date = (row.get("startDate") or "").strip()
    end_date = (row.get("endDate") or "").strip()
    if ical:
        rec: Dict[str, str] = {"icalString": ical}
        if start_date:
            rec["startDate"] = start_date
        if end_date:
            rec["endDate"] = end_date
        return {"recurrences": [rec]}
    return {}

def is_on_demand(schedule: Dict) -> bool:
    #Check if job should run immediately (on-demand).
    return not schedule

def load_accounts_from_env() -> dict:
    """
    Build {login_email_lower: {"baseurl":..., "username":..., "password":...,}} from .env only.
    Optional password-file override (scheduler only):
      FUSION_SCHEDULER_LOGIN=<login to override>
      -> if set and ./password/password.txt exists/non-empty, that password replaces the .env password.
      SPECIAL FEATURE - Password File Override:
        If FUSION_SCHEDULER_LOGIN is set, and ./password/password.txt exists,
        the password file overrides the .env password for that login.
        This is where it shows that can have txt or just env for passwords
        
        This is useful for rotating scheduler passwords without editing .env
        
    """
    base = (os.getenv("FUSION_BASEURL") or "").rstrip("/")
    aliases = [a.strip() for a in (os.getenv("FUSION_USERS") or "").split(",") if a.strip()]
    if not base or not aliases:
        raise SystemExit("Missing FUSION_BASEURL or FUSION_USERS in .env")

    scheduler_login_env = (os.getenv("FUSION_SCHEDULER_LOGIN") or "").strip().lower()
    out = {}

    for alias in aliases:
        login = (os.getenv(f"FUSION_{alias}_LOGIN") or "").strip()
        pwd   = (os.getenv(f"FUSION_{alias}_PASSWORD") or "").strip()
        if not login:
            raise SystemExit(f"Missing login for alias {alias}: set FUSION_{alias}_LOGIN")

        # Apply override if this alias matches scheduler login
        if scheduler_login_env and login.lower() == scheduler_login_env:
            file_pwd = _read_password_file()
            if file_pwd:
                pwd = file_pwd
                logger.info(f"[auth] Using password file for {login} (source=file)")
            else:
                logger.info(f"[auth] Using .env password for {login} (source=env)")

        if not pwd:
            raise SystemExit(
                f"Missing password for alias {alias}: set FUSION_{alias}_PASSWORD "
                f"or provide password/password.txt if this alias is the scheduler."
            )

        out[login.lower()] = {"baseurl": base, "username": login, "password": pwd}
    return out

def pick_login_from_row(row: dict) -> str:
    #requires username column with actual username
    #extacts username from column reason is so I can run code through multiple users, say one job has specific privilige assigned to user1 and user2 does not have role,
    for k in row.keys():
        if k and k.strip().lower() == "username":
            v = (row.get(k) or "").strip()
            if not v:
                raise ValueError("Excel row missing Username login")
            return v
    raise ValueError("Excel is missing a 'Username' column header")

def build_client_cache_env_only():
    """
    Build a client getter function with caching.
    
    WHAT IT DOES:
        1. Load all accounts from environment variables
        2. Return a function that creates/caches ESSClient instances
        3. Reuse clients for the same username (connection pooling)
    
    RETURNS:
        Function: get(login_email: str) -> ESSClient
    
    HOW IT WORKS:
        First call for a login: Creates new ESSClient, caches it
        Subsequent calls: Returns cached client (reuses connection)
    
    DESIGN DECISION - Lazy Client Creation:
        Clients only created when actually needed
        If Excel only uses one account, only one client is created
        Saves authentication overhead for unused accounts
    
    DESIGN DECISION - Connection Reuse:
        ESSClient sessions are reused across multiple jobs
        Reduces authentication overhead (login once, submit many jobs)
    Was tyring to mae code faster. So don't have to login.
        """
    accounts_by_login = load_accounts_from_env()
    cache = {}
    def get(login_email: str):
        """Get or create ESSClient for given login."""
        key = login_email.lower().strip()
        if key not in accounts_by_login:
            known = ", ".join(accounts_by_login.keys())
            raise ValueError(f"Unknown Username '{login_email}'. Known: {known}")
        if key not in cache:
            cfg = accounts_by_login[key]
            cache[key] = ESSClient(cfg["baseurl"], cfg["username"], cfg["password"])
        return cache[key]
    return get
# ============================================================================
# STATUS TRACKER HELPERS
# ============================================================================

# ==== ACL job identifiers (CRITICAL jobs - exact names only) ====
ACL_JOB_NAMES = [
    "Compute Users ACL",
    "Compute Users ACL by Event",
    "Compute Users with Large ACL",
      # Add 4th job name here if different
]

def is_acl_job(job_name: str) -> bool:
    """
    Check if job is one of the 4 critical ACL jobs.
    
    WHAT IT DOES:
        Checks if job name exactly matches one of the 4 ACL job names
        (case-insensitive comparison)
    
    CRITICAL ACL JOBS:
        1. Compute Users ACL
        2. Compute Users ACL by Event
        3. Compute Users with Large ACL
        4. Compute Users ACL
    
    EXAMPLES:
        "Compute Users ACL" → True (exact match)
        "compute users acl" → True (case-insensitive)
        "Compute Users ACL - Test" → False (not exact)
        "Import Payables" → False (not an ACL job)
    """
    job_lower = job_name.strip().lower()
    return any(acl_job.lower() == job_lower for acl_job in ACL_JOB_NAMES)


def init_status_tracker() -> dict:
    """
    Initialize empty status tracker structure.
    
    RETURNS:
        Empty status dict with all fields set to defaults
    """
    return {
        "rest_api_completed": False,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "total_jobs": 0,
        "successful": 0,
        "failed": 0,
        "timed_out": 0,
        "skipped": 0,
        "failed_jobs": [],
        "acl_jobs_status": {
            "attempted": False,
            "completed": True,  # Assume success unless we find failures
            "failed_acl_jobs": []
        }
    }


def write_status_tracker(status_data: dict):
    """
    Write status tracker to JSON file.
    
    WHAT IT DOES:
        Writes the status dictionary to job_status_tracker.json
        This file is read by ui_ess_jobs.py to know which jobs to retry
    
    PARAMETERS:
        status_data: Complete status dictionary
    """
    try:
        with open(STATUS_FILE, 'w', encoding='utf-8') as f:
            json.dump(status_data, f, indent=2, ensure_ascii=False)
        logger.info(f" Status tracker written to {STATUS_FILE}")
    except Exception as e:
        logger.error(f" Failed to write status tracker: {e}")


def add_failed_job(status_data: dict, display_name: str, request_id: str,
                   reason: str, row_num: int, params: dict):
    """
    Add a failed job to the status tracker.
    
    WHAT IT DOES:
        1. Adds job details to failed_jobs list
        2. If it's an ACL job, adds it to critical failed list
        3. Logs a warning (or error if ACL job)
    
    PARAMETERS:
        status_data: The status tracker dict
        job_name: Display name or job definition name
        request_id: Oracle request ID (empty string if not submitted)
        reason: Why it failed (TIMEOUT_WAIT, TIMEOUT_OVERALL, FAILED, etc.)
        row_num: Excel row number (for debugging)
        params: Job parameters dict
    """
    

    failed_job = {
       "job_name": display_name,
        "request_id": str(request_id),
        "reason": reason,
        "excel_row": row_num,
        "parameters": params
    }
    
    status_data["failed_jobs"].append(failed_job)
    
    # Check if ACL job (CRITICAL)
    if is_acl_job(display_name):
        status_data["acl_jobs_status"]["attempted"] = True
        status_data["acl_jobs_status"]["completed"] = False
        status_data["acl_jobs_status"]["failed_acl_jobs"].append(display_name)
        logger.error(f" CRITICAL: ACL JOB FAILED: {display_name} (Reason: {reason})")
    else:
        logger.warning(f" Job failed: {display_name} (Reason: {reason})")
# ------------------------- MAIN RUN FUNCTION -------------------------
def run() -> int:
    """
    Main entry point - process Excel file(s) and submit ESS jobs.
    
    WORKFLOW:
        1. Parse command line arguments (--scenario, --folder, --strict-order)
        2. Collect Excel file paths to process
        3. Build client cache from environment variables
        4. Initialize status tracker for failed job tracking
        5. For each Excel file:
           a. Stage file as data.xlsx (if not already)
           b. Read rows from Excel
           c. For each row:
              - Get username and create/reuse ESSClient
              - Extract job identifiers (ID, package, name, application)
              - Resolve job by display name if needed
              - Extract parameters and schedule
              - Check for duplicate running jobs
              - Submit job with timeout-aware polling
              - Track success/timeout/failure in status tracker
              - Log results to Excel audit file
        6. Write final status tracker to JSON
        7. Return error code if failures detected
    
    MULTI-FILE SUPPORT:
        Scenario mode: Process predefined file list (e.g., (client)P2T.xlsx)
        Folder mode: Process all Excel files in scenarios/<folder>/
        Default mode: Process single data.xlsx from drop folder
    
    MULTI-TENANT SUPPORT:
        Each row can specify different Username
        ESSClient instances cached per username
        One authentication session reused for multiple jobs
    
    TIMEOUT TRACKING:
        On-demand jobs monitored for:
        - Overall timeout (MAX_POLL_TIME_SECONDS)
        - WAIT state timeout (MAX_WAIT_STATE_TIME)
        Failed/timed out jobs written to job_status_tracker.json
    
    ERROR HANDLING:
        File-level errors: Log and skip to next file
        Row-level errors: Log, track in status, continue to next row
        Returns 1 if any jobs failed/timed out (triggers UI retry)
    
    AUDIT LOGGING:
        Every submission logged to logs/job_runs.xlsx
        Separate entries for on-demand vs scheduled jobs
        Tracks: timestamp, username, job name, request ID, final status
    
    RETURNS:
        0 on success (all jobs succeeded)
        1 on partial failure (some jobs failed/timed out - triggers UI retry)
    """
    # Parse arguments
    args = _parse_args()

    # Collect files to process
    excel_paths = _collect_excel_paths(args)
    if not excel_paths:
        logger.info("No Excel files found to process.")
        return 0

    get_client = build_client_cache_env_only()
    
    # Initialize status tracker
    status_tracker = init_status_tracker()
    
    # Counters
    total_rows = 0
    success_count = 0
    timeout_count = 0
    failed_count = 0
    skipped_count = 0
    pending_jobs = []  # list of dicts: {request_id, display_name, job_name_for_tracking, row_num, login, params, client}
    for xlsx in excel_paths:
        try:
            # Stage each Excel as data.XLSX
            if xlsx.resolve() != DATA_XLSX_PATH.resolve():
                if not _is_file_stable(xlsx, DATA_STABLE_SEC):
                    logger.warning(f"File not yet stable (skip this run): {xlsx.name}")
                    continue
                shutil.copy2(xlsx, DATA_XLSX_PATH)
                logger.info(f"Using {xlsx.name} → staged as {DATA_XLSX_PATH.name}")
            else:
                logger.info(f"Using existing {DATA_XLSX_PATH.name}.")

            rows = read_jobs_from_excel(DATA_XLSX_PATH)
            if not rows:
                logger.warning(f"No rows found in {xlsx.name}")
                continue

            for i, row in enumerate(rows, start=1):
                # ── 4A: Enabled column check ──
                enabled = str(row.get("Enabled", "Y")).strip().upper()
                if enabled in ("N", "NO", "FALSE", "0", ""):
                    logger.info(f"Row {i}: SKIPPED (Enabled={row.get('Enabled', '')})")
                    continue
                login = pick_login_from_row(row)
                client = get_client(login)
                logger.info(f"[{login}] {xlsx.name} row {i} selected account")
                total_rows += 1  # count every row attempted regardless of outcome

                job_definition_id   = (row.get("jobDefinitionId") or "").strip()
                job_package_name    = (row.get("jobPackageName") or "").strip().rstrip("/")
                job_definition_name = (row.get("jobDefinitionName") or "").strip()
                application         = (row.get("application") or "").strip()
                description         = (row.get("Display Name") or "").strip()
                class_name          = (row.get("className") or "").strip()
                display_name        = description
                
                # Job name for tracking (prefer jobDefinitionName, fallback to Display Name)
                job_name_for_tracking = job_definition_name or display_name or "UnknownJob"

                # Resolve job by display name if necessary
                if (not job_definition_id and not (job_package_name and job_definition_name)) or not application:
                    try:
                        resolved = client.resolve_job_by_name(display_name)
                    except Exception:
                        resolved = None
                    if resolved:
                        application = application or resolved.get("application") or application
                        if not (job_package_name and job_definition_name):
                            job_package_name    = job_package_name or resolved.get("jobPackageName") or ""
                            job_definition_name = job_definition_name or resolved.get("jobDefinitionName") or ""
                            job_definition_id   = job_definition_id or resolved.get("jobDefinitionId") or job_definition_id
                            # Update tracking name if we resolved it
                            job_name_for_tracking = job_definition_name or display_name or "UnknownJob"
                        logger.info(
                            f"Resolved '{display_name}' → app={application or '-'} "
                            f"pkg={job_package_name or '-'} name={job_definition_name or '-'}"
                        )
                    else:
                        logger.error(
                            f"Row {i}: cannot resolve jobDefinition fields from Display Name '{display_name}'. Skipping."
                        )
                        failed_count += 1
                        add_failed_job(
                            status_tracker,
                            job_name_for_tracking,
                            "",
                            "RESOLUTION_FAILED",
                            i,
                            {}
                        )
                        continue

                # Prefer jobDefinitionId path over package+name
                if job_package_name and job_definition_name and not job_definition_id:
                    job_definition_id = f"JobDefinition://{job_package_name.lstrip('/')}/{job_definition_name}"
                if job_definition_id:
                    job_package_name = ""
                    job_definition_name = ""

                # Extract parameters
                params = extract_params(row)

                # Mirror generic args onto submit.argumentN (and vice versa)
                for n in range(1, 30 + 1):
                    a  = f"argument{n}"
                    sa = f"submit.argument{n}"
                    if params.get(a) and not params.get(sa):
                        params[sa] = params[a]
                    if params.get(sa) and not params.get(a):
                        params[a] = params[sa]

                # OSCS alias fan-out
                def _is_oscs_job() -> bool:
                    jdid = (job_definition_id or "").lower()
                    pkg  = (job_package_name or "").lower()
                    name = (job_definition_name or "").lower()
                    disp = (display_name or "").lower()
                    return (
                        "oscs" in disp
                        or "fndoscs" in name
                        or "/fnd/applcore/" in jdid
                        or "/fnd/applcore/" in pkg
                    )

                if _is_oscs_job():
                    OSCS_PARAM_ALIASES = [
                        "Index Name to Reingest",
                        "Index Name for recreate",
                        "indexName",
                        "argument1",
                        "submit.argument1",
                    ]
                    ix_val = next((params.get(k) for k in OSCS_PARAM_ALIASES if params.get(k)), None)
                    if ix_val:
                        for key in ("Index Name for recreate", "indexName", "argument1", "submit.argument1"):
                            params.setdefault(key, ix_val)

                if not params:
                    logger.warning(
                        f"Row {i}: no parameters extracted for '{display_name}'. "
                        f"If the job requires prompts, ensure Excel headers match exactly "
                        f"(e.g., 'Business Unit', 'Ledger') or use 'argument1..30'."
                    )

                # Schedule extraction / validation
                try:
                    schedule = extract_schedule(row)
                except Exception as e:
                    logger.exception(f"Row {i}: invalid schedule: {e}")
                    failed_count += 1
                    add_failed_job(
                        status_tracker,
                        job_name_for_tracking,
                        "",
                        "INVALID_SCHEDULE",
                        i,
                        params
                    )
                    continue

                logger.info(f"Row {i}: parameters (dict) → {json.dumps(params, ensure_ascii=False)}")

                # Duplicate / existing handling
                allow_dup = str(row.get("AllowDuplicate", "")).strip().upper() in ("Y","YES","TRUE","1")   # <-- add
                existing_id = None
                row_start_date = (row.get("startDate") or "").strip()
                row_argument1  = params.get("argument1", "").strip()
                if job_definition_id and not allow_dup:
                    try:
                        if is_on_demand(schedule):
                            # Pass argument1 so OSCS jobs with different indexes
                            # are not blocked by a same-definition running job
                            existing_id = client.find_active_request_for_definition(
                                job_definition_id,
                                argument1=row_argument1,
                            )
                        else:
                            # Scheduled jobs: match startDate + argument1 too
                            existing_id = client.find_active_scheduled_request(
                                job_definition_id,
                                start_date=row_start_date,
                                argument1=row_argument1,
                            )
                    except Exception:
                        existing_id = None
                if existing_id:
                    logger.info(
                        f"Existing active request for {job_definition_id}: request_id={existing_id} — not submitting a duplicate."
                    )
                    if is_on_demand(schedule):
                        if DRY_RUN:
                            logger.info("[DRY RUN] Would poll the existing request to completion (skipped).")
                        else:
                            final = client.poll_until_complete(
                                existing_id,
                                poll_seconds=int(os.getenv("DUPLICATE_POLL_INTERVAL_SECONDS", "10")),
                                timeout_seconds=int(os.getenv("DUPLICATE_POLL_TIMEOUT_SECONDS", "300")),
                                logger=logger,
                            )
                            final_status = final.get("Status") or final.get("status") or final
                            logger.info(
                                f"[ON-DEMAND EXISTING] request_id={existing_id} def={job_definition_id} final_status={final_status}"
                            )
                            # Track result
                            # If job is still running/queued after the short poll window,
                            # that means Oracle is processing it normally — not a failure.
                            # Only flag as failed if Oracle returned a true error state.
                            final_status_upper = str(final_status).upper()
                            STILL_RUNNING_STATES = {"RUNNING", "WAIT", "BLOCKED", "HOLD", "PAUSED", "PENDING", "READY", "UNKNOWN"}
                            if final_status_upper in ("SUCCEEDED", "SUCCESS"):
                                success_count += 1
                            elif final_status_upper in STILL_RUNNING_STATES:
                                skipped_count += 1
                                logger.info(
                                    f"[ON-DEMAND EXISTING] request_id={existing_id} job={display_name} "
                                    f"still {final_status_upper} after poll window — in-flight, treating as skipped"
                                )
                            else:
                                # Genuine failure: ERROR, FAILED, WARNING, CANCELLED
                                failed_count += 1
                                add_failed_job(
                                    status_tracker,
                                    display_name,
                                    existing_id,
                                    f"EXISTING_{final_status_upper}",
                                    i,
                                    params
                                )
                            # Audit log
                            try:
                                log_job_run_xlsx(
                                    job_name=display_name,  # also changed to match
                                    request_id=str(existing_id),
                                    status=str(final_status),
                                    fusion_alias=str(login),
                                    notes="existing_on_demand"
                                )
                            except Exception:
                                pass
                    else:
                        logger.info(f"[SCHEDULED] Another schedule/run already active for def={job_definition_id}; skipping submit.")
                        skipped_count += 1
                    continue
                # Submit new job
                try:
                    req_id = client.submit(
                        job_definition_id,
                        application,
                        params,
                        schedule if not is_on_demand(schedule) else {},
                        description=description,
                        dry_run=DRY_RUN,
                        logger=logger,
                        job_package_name=job_package_name or None,
                        job_definition_name=job_definition_name or None,
                        class_name=class_name or None,
                    )

                    if DRY_RUN:
                        if is_on_demand(schedule):
                            logger.info("[DRY RUN] Would poll until completion for on-demand job (skipped).")
                        else:
                            logger.info("[DRY RUN] Would log scheduled parent requestId (skipped).")
                        continue

                    # ========== PHASE 1: COLLECT FOR BATCH POLLING ==========
                    if is_on_demand(schedule):
                        logger.info(f"  Submitted on-demand job: request_id={req_id} — will poll after all submissions")
                        pending_jobs.append({
                            "request_id": req_id,
                            "display_name": display_name,
                            "job_name_for_tracking": job_name_for_tracking,
                            "row_num": i,
                            "login": login,
                            "params": params,
                            "client": client,
                        })
                        time.sleep(10)
                    
                    
                    # ========== SCHEDULED JOBS ==========
                    else:
                        success_count += 1
                        logger.info(
                            f"[SCHEDULED] parent_request_id={req_id} job={job_name_for_tracking} schedule_created"
                        )
                        # Audit log
                        try:
                            log_job_run_xlsx(
                                job_name=job_name_for_tracking,
                                request_id=str(req_id),
                                status="SCHEDULED",
                                fusion_alias=str(login),
                                notes="scheduled_parent"
                            )
                        except Exception:
                            pass

                except ESSDuplicatePendingError:
                    logger.warning(f"ESS-01050: another request is already pending for {job_definition_id}.")
                    skipped_count += 1
                    continue
                except Exception as e:
                    logger.exception(f"Row {i}: submit failed: {e}")
                    failed_count += 1
                    add_failed_job(
                        status_tracker,
                        job_name_for_tracking,
                        "",
                        f"EXCEPTION: {str(e)}",
                        i,
                        params
                    )

        except Exception as e:
            logger.exception(f"Failed processing file: {xlsx.name} → {e}")
# ========== PHASE 2: BATCH POLL ALL SUBMITTED ON-DEMAND JOBS ==========
    if pending_jobs and not DRY_RUN:
        GLOBAL_POLL_TIMEOUT = int(os.getenv("ESS_POLL_TIMEOUT", "1800"))  # 30 minutes max for all jobs combined 1800
        TERMINAL_STATES = {"SUCCEEDED", "SUCCESS", "ERROR", "FAILED", "WARNING", "CANCELLED"}
        logger.info(f"\n{'='*60}")
        logger.info(f" PHASE 2: POLLING {len(pending_jobs)} SUBMITTED JOBS")
        logger.info(f" Global timeout: {GLOBAL_POLL_TIMEOUT}s ({GLOBAL_POLL_TIMEOUT//60} minutes)")
        logger.info(f"{'='*60}\n")
        
        poll_global_start = time.time()
        outstanding = list(pending_jobs)  # copy — we'll remove as they complete
        
        while outstanding:
            global_elapsed = int(time.time() - poll_global_start)
            
            # Global timeout check
            if global_elapsed > GLOBAL_POLL_TIMEOUT:
                logger.error(f" Global poll timeout reached ({global_elapsed}s). {len(outstanding)} jobs still outstanding.")
                for job in outstanding:
                    timeout_count += 1
                    job["final_status"] = "TIMEOUT_GLOBAL"
                    add_failed_job(
                        status_tracker,
                        job["display_name"],
                        job["request_id"],
                        "TIMEOUT_GLOBAL",
                        job["row_num"],
                        job["params"]
                    )
                    try:
                        log_job_run_xlsx(
                            job_name=job["job_name_for_tracking"],
                            request_id=str(job["request_id"]),
                            status="TIMEOUT_GLOBAL",
                            fusion_alias=str(job["login"]),
                            notes=f"global_timeout_after_{global_elapsed}s"
                        )
                    except Exception:
                        pass
                break
            
            # Poll each outstanding job once
            still_outstanding = []
            for job in outstanding:
                try:
                    status_data = job["client"].get_status(job["request_id"])
                    current_state = (status_data.get("Status") or status_data.get("status") or "UNKNOWN").upper()
                except Exception as e:
                    logger.warning(f"  Failed to get status for {job['request_id']}: {e}")
                    still_outstanding.append(job)
                    continue
                
                if current_state in TERMINAL_STATES:
                    # Job finished
                    elapsed = int(time.time() - poll_global_start)
                    logger.info(f"  {job['display_name']} (req={job['request_id']}) → {current_state} ({elapsed}s)")
                    job["final_status"] = current_state          
                    job["elapsed"] = elapsed
                    if current_state in ("SUCCEEDED", "SUCCESS"):
                        success_count += 1
                    else:
                        failed_count += 1
                        add_failed_job(
                            status_tracker,
                            job["display_name"],
                            job["request_id"],
                            current_state,
                            job["row_num"],
                            job["params"]
                        )
                    
                    try:
                        log_job_run_xlsx(
                            job_name=job["job_name_for_tracking"],
                            request_id=str(job["request_id"]),
                            status=current_state,
                            fusion_alias=str(job["login"]),
                            notes=f"batch_poll_{elapsed}s"
                        )
                    except Exception:
                        pass
                else:
                    # Still running
                    still_outstanding.append(job)
            
            outstanding = still_outstanding
            
            if outstanding:
                logger.info(f"  [{global_elapsed}s] {len(outstanding)} jobs still running, {len(pending_jobs) - len(outstanding)} completed. Waiting {POLL_INTERVAL_SECONDS}s...")
                time.sleep(POLL_INTERVAL_SECONDS)
        
        logger.info(f"\n{'='*60}")
        logger.info(f" PHASE 2 COMPLETE: All jobs reached terminal state")
        logger.info(f"{'='*60}\n")
        # ── Per-job Process ID table ──
        logger.info(f"\n{'='*72}")
        logger.info(f" ON-DEMAND JOB DETAIL (Process IDs)")
        logger.info(f"{'='*72}")
        logger.info(f" {'Process ID':<12} {'Status':<16} {'Arg1':<22} {'Job'}")
        logger.info(f" {'-'*11:<12} {'-'*15:<16} {'-'*21:<22} {'-'*20}")
        for job in pending_jobs:
            pid   = str(job.get("request_id", "") or "")
            stat  = job.get("final_status", "UNKNOWN")
            arg1  = str((job.get("params") or {}).get("argument1", ""))[:20]
            name  = str(job.get("display_name", ""))[:45]
            logger.info(f" {pid:<12} {stat:<16} {arg1:<22} {name}")
        logger.info(f"{'='*72}\n")
    
    elif pending_jobs and DRY_RUN:
        logger.info(f"[DRY RUN] Would batch-poll {len(pending_jobs)} on-demand jobs (skipped).")
    # ========== FINALIZE STATUS TRACKER ==========
    status_tracker["rest_api_completed"] = True
    status_tracker["timestamp"] = datetime.now(timezone.utc).isoformat()
    status_tracker["total_jobs"] = total_rows
    status_tracker["successful"] = success_count
    status_tracker["failed"] = failed_count
    status_tracker["timed_out"] = timeout_count
    status_tracker["skipped"] = skipped_count
    
    # Write to file
    write_status_tracker(status_tracker)
    
    # ========== SUMMARY ==========
    logger.info(f"\n{'='*60}")
    logger.info(f" REST API JOB SUMMARY")
    logger.info(f"{'='*60}")
    logger.info(f" Successful:  {success_count}")
    logger.info(f"  Timed out:   {timeout_count}")
    logger.info(f" Failed:      {failed_count}")
    logger.info(f" Skipped:     {skipped_count}")
    logger.info(f" Total:       {total_rows}")
    
    # ACL warning
    if not status_tracker["acl_jobs_status"]["completed"]:
        logger.error(f"\n{'🚨'*30}")
        logger.error(f"🚨 CRITICAL: ACL JOBS FAILED!")
        logger.error(f"Failed ACL jobs: {status_tracker['acl_jobs_status']['failed_acl_jobs']}")
        logger.error(f"{'🚨'*30}\n")
    
    logger.info(f"{'='*60}")
    logger.info(f"Processed {total_rows} rows across {len(excel_paths)} file(s).")
    
    # Return code: 0 = all success, 1 = some failures (triggers UI retry)
    if timeout_count + failed_count > 0:
        logger.warning(f"  {timeout_count + failed_count} jobs need UI retry")
        return 1
    
    return 0


if __name__ == "__main__":
    # Check if a scenario argument (like "P2T") was passed
    # If not, it defaults to None (which falls back to data.XLSX)
    target_scenario = None
    if len(sys.argv) > 1:
        target_scenario = sys.argv[1]
        
    # create a fake "args" object so I don't have to rewrite collect_excel_paths function
    class SimpleArgs:
        def __init__(self, scenario, folder=None, strict_order=False):
            self.scenario = scenario
            self.folder = folder
            self.strict_order = strict_order

    # If UI script sent "P2T", this will trigger Scenario Mode
    # and look for (client)P2T.xlsx in your folder
    print(f"Launching REST API Logic for Scenario: {target_scenario or 'Default'}")
    
    # Run the main process
    raise SystemExit(run())