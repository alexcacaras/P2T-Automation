from __future__ import annotations
# ==========================================================================
# backfill_from_log.py - Parse ESS Logs into Excel Audit Trail
# ============================================================================
"""
WHAT THIS FILE DOES:
    Parses ess.log text files and extracts job execution records into an Excel
    audit trail (logs/job_runs.xlsx, "Runs_v2" sheet). This provides historical
    visibility into job runs that may not have been captured by the real-time
    Excel logging system.

WHY THIS EXISTS:
    The main automation (main.py) logs to Excel in real-time, but:
    1. If Excel file is locked, some runs may be missed
    2. Historical logs exist but aren't in Excel format
    3. Need to reconstruct audit trail after log rotation
    4. Want unified view combining real-time and backfilled data
    5.Live Excel log fails
    6.In short I wanted way to log the log files in more readable format

USAGE:
    # Backfill all historical logs:
    python backfill_from_log.py
    
    # Or import and call:
    from backfill_from_log import backfill
    backfill()

LOG PARSING STRATEGY:
    Uses regex patterns to extract job metadata from log lines:
    - Timestamps (ISO 8601 UTC)
    - Fusion alias ([username])
    - Display names ("Submitting job X", "Resolved 'X'")
    - Job identifiers (name, ID, package, application)
    - Request IDs (request_id=12345)
    - Final status (SUCCEEDED, FAILED, etc.)
    - Error snippets (ESS-xxxx, ORA-xxxx, Exception)

KEY FEATURES:
    - Request-ID-aware parsing (merges multi-line job info)
    - Handles both free-text and JSON-formatted log lines
    - UTC → Local timezone conversion
    - Time window filtering (last N days)
    - Shadow file fallback (if Excel locked)
    - Progress reporting every 10 rows

CONFIGURATION:
    LOG_PATHS: List of log files to parse (default: logs/ess.log)
    LAST_N_DAYS: Filter to recent days (None = all history)
    SHEET_NAME: Excel sheet name (default: "Runs_v2")
    LOCAL_TZ: Timezone for local_date/local_time columns (default: America/Toronto)

OUTPUT EXCEL COLUMNS:
    - timestamp_utc: ISO timestamp in UTC
    - local_date: Date in local timezone
    - local_time: Time in local timezone
    - fusion_alias: Username who submitted job
    - display_name: Human-readable job name
    - job_definition_name: Short job name
    - job_definition_id: Full job path
    - job_package_name: Job package
    - application: Oracle module (ERP, HCM, etc.)
    - request_id: Oracle request ID
    - status: SUCCEEDED, FAILED, DRY_RUN, etc.
    - error: Short error snippet (if any)
    - notes: "backfilled" marker

DEPENDENCIES:
    - openpyxl: Excel file reading/writing
    - zoneinfo: Timezone conversion (Python 3.9+)

KNOWN LIMITATIONS:
    - Only extracts info present in log lines
    - May miss jobs if log format changes
    - Regex-based parsing (fragile to log format changes)
    - No deduplication (same job may appear multiple times)
"""
import re, sys   #re for regular expressions, sys for runtime utilities 
from datetime import datetime, timedelta, timezone #for timestamps and different timezones
from pathlib import Path #for paths
from typing import Iterable, Dict, Optional #for type hints

from openpyxl import Workbook, load_workbook #for creating/loading excel workbooks
from openpyxl.utils import get_column_letter #for autozing columns 

# ---------------- Config ----------------
LOG_PATHS = [Path("logs") / "ess.log"]   # add rotated logs if needed
LAST_N_DAYS: Optional[int] = None        # e.g., 7 for last week; None = all history
PROGRESS_EVERY = 10
SHEET_NAME = "Runs_v2"
HEADERS = [
    "timestamp_utc", "local_date", "local_time",
    "fusion_alias",
    "display_name", "job_definition_name", "job_definition_id", "job_package_name", "application",
    "request_id", "status", "error", "notes",
]
LOCAL_TZ = "America/Toronto"             # change if you want a different local view
# ---------------------------------------

def _default_xlsx_path(): #get default path
    base = (Path.cwd() / "logs")
    base.mkdir(parents=True, exist_ok=True) #make directory
    return base / "job_runs.xlsx"

def _autosize(ws): #auto size columns and make 60chars cap for redability
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len: max_len = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def _ensure_wb_sheet(path: Path):
    # try to open; if locked, write to shadow file
    def _new(path_): #create new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        _autosize(ws)
        wb.save(path_)
        return wb, ws, path_

    if not path.exists():
        return _new(path)

    try:
        wb = load_workbook(path)
        real_path = path
    except PermissionError:
        #file locked use shadow file
        real_path = path.with_name(f"{path.stem}__shadow{path.suffix}")
        if not real_path.exists():
            return _new(real_path)
        wb = load_workbook(real_path)

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    # ensure correct header
    if ws.max_row == 0:
        ws.append(HEADERS)
    else:
        row1 = [ws.cell(row=1, column=i+1).value for i in range(len(HEADERS))]
        if row1 != HEADERS:
            ws = wb.create_sheet(f"{SHEET_NAME}_{len(wb.sheetnames)}")
            ws.append(HEADERS)
    return wb, ws, real_path

def _local_parts(ts_utc_iso: str, tz_name: str = LOCAL_TZ): #needed way to better read timezone
    try:
        import zoneinfo #for converting utc to local timezone
        dt = datetime.strptime(ts_utc_iso, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
        lt = dt.astimezone(zoneinfo.ZoneInfo(tz_name))
        return lt.strftime("%Y-%m-%d"), lt.strftime("%H:%M:%S")
    except Exception:
        return "", ""

# --------- parsing regex (covers both free-text and JSON-like keys) ---------
TS_RE         = re.compile(r"(?P<ts>\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z)")
ALIAS_RE      = re.compile(r"\[(?P<alias>[^\]]+)\]")

# display name sources
SUBMIT_RE     = re.compile(r"Submitting\s+job\s+(?P<disp>.+)", re.I)
RESOLVED_RE   = re.compile(r"Resolved\s+'(?P<disp>[^']+)'", re.I)

# definition name / id / package / app (free text)
NAME_RE       = re.compile(r"\bname=(?P<name>[\w\.]+)\b", re.I)
JDID_TXT_RE   = re.compile(r"(JobDefinition://[^\s,;]+)")
PKG_TXT_RE    = re.compile(r"\bpkg\s*=\s*(?P<pkg>[^,\s]+)", re.I)
APP_TXT_RE    = re.compile(r"\bapp(?:lication)?\s*=\s*(?P<app>[^,\s]+)", re.I)

# same from JSON-like lines:
PKG_JSON_RE   = re.compile(r'"jobPackageName"\s*:\s*"(?P<pkg>[^"]+)"', re.I)
NAME_JSON_RE  = re.compile(r'"jobDefinitionName"\s*:\s*"(?P<name>[^"]+)"', re.I)
APP_JSON_RE   = re.compile(r'"application"\s*:\s*"(?P<app>[^"]+)"', re.I)

RID_RE        = re.compile(r"request[_ ]?id\s*=\s*(?P<rid>\d+)", re.I)
FINAL_WORD    = re.compile(r"\b(SUCCEEDED|FAILED|ERROR|WARNING|CANCELED)\b", re.I)
FINAL_KV      = re.compile(r"\bfinal_status\s*=\s*(?P<status>[A-Z_]+)", re.I)
DRY_RE        = re.compile(r"\bDRY[_ ]?RUN\b", re.I)

ERROR_LINE_RE = re.compile(r"(ESS-\d+|ORA-\d+|JBO-\d+|ADF-\d+|FND-\d+|SVC-\d+|BEA-\d+|ERROR|Exception|Traceback)", re.I)

_cutoff = None if LAST_N_DAYS is None else datetime.now(timezone.utc) - timedelta(days=LAST_N_DAYS)
def _in_window(ts_iso: str) -> bool: #Check if timestamp is within configured time window so only get from tiemframe
    if _cutoff is None: return True
    try:
        dt = datetime.strptime(ts_iso, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
        return dt >= _cutoff
    except Exception:
        return True

def _line_ts(line: str, fallback: Optional[str]) -> str: #get timestamp
    m = TS_RE.search(line)
    if m: return m.group("ts")
    return fallback or datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def _cut(text: str, n: int = 220) -> str: #trunc to max length with ... to indicate if more text
    t = (text or "").strip()
    return (t[:n] + "…") if len(t) > n else t

def parse_log(lines: Iterable[str]) -> Iterable[Dict[str, str]]:
    """
    Parse log lines into job execution records.
    
    PARSING STRATEGY:
        Uses two context dicts:
        1. current: Accumulates info for job being processed
        2. by_rid: Accumulates info by request_id (for multi-line jobs)
    
    WHAT TRIGGERS OUTPUT (YIELD):
        - Line contains "DRY RUN" → Emit DRY_RUN status
        - Line contains final status word → Emit with that status
    
    REQUEST-ID-AWARE BEHAVIOR:
        When request_id found, all subsequent info for that ID gets merged
        Example:
            Line 1: "Submitting job X request_id=123"
            Line 2: "request_id=123 → SUCCEEDED"
            Result: Single record with job name + final status
    
    DUAL FORMAT SUPPORT:
        Extracts from both:
        - Free-text: "pkg=/oracle/apps name=MyJob app=ERP"
        - JSON: '{"jobPackageName":"/oracle/apps","jobDefinitionName":"MyJob"}'
    
    ERROR CAPTURE:
        If line contains error pattern (ESS-xxxx, Exception, etc.)
        Captures snippet in "error" field (truncated to 220 chars)
    
    YIELDS:
        Dict with keys matching HEADERS list
        One dict per completed job run found in logs
        """
    current: Dict[str, str] = {}
    by_rid: Dict[str, Dict[str, str]] = {}

    def upsert(ctx: Dict[str, str], line: str):
        if (m := TS_RE.search(line)) and "timestamp_utc" not in ctx:
            ctx["timestamp_utc"] = m.group("ts")
        if (m := ALIAS_RE.search(line)):
            ctx.setdefault("fusion_alias", m.group("alias"))

        # display name
        if (m := SUBMIT_RE.search(line)):
            ctx["display_name"] = m.group("disp").strip()
        if (m := RESOLVED_RE.search(line)) and "display_name" not in ctx:
            ctx["display_name"] = m.group("disp").strip()

        # def name / id / pkg / app (free-text)
        if (m := NAME_RE.search(line)) and "job_definition_name" not in ctx:
            ctx["job_definition_name"] = m.group("name").strip()
        if (m := JDID_TXT_RE.search(line)) and "job_definition_id" not in ctx:
            ctx["job_definition_id"] = m.group(1).strip()
        if (m := PKG_TXT_RE.search(line)) and "job_package_name" not in ctx:
            ctx["job_package_name"] = m.group("pkg").strip()
        if (m := APP_TXT_RE.search(line)) and "application" not in ctx:
            ctx["application"] = m.group("app").strip()

        # def name / pkg / app (JSON-like)
        if (m := NAME_JSON_RE.search(line)):  ctx.setdefault("job_definition_name", m.group("name").strip())
        if (m := PKG_JSON_RE.search(line)):   ctx.setdefault("job_package_name",   m.group("pkg").strip())
        if (m := APP_JSON_RE.search(line)):   ctx.setdefault("application",        m.group("app").strip())

        # an error-looking line → capture short snippet
        if ERROR_LINE_RE.search(line):
            ctx["error"] = _cut(line)

        # request id
        if (m := RID_RE.search(line)):
            rid = m.group("rid")
            ctx["request_id"] = rid
            store = by_rid.get(rid, {})
            store.update(ctx)
            by_rid[rid] = store

    for raw in lines:
        line = raw.rstrip("\n")
        if not line.strip():
            continue

        upsert(current, line)

        m_rid = RID_RE.search(line)
        ctx = by_rid.get(m_rid.group("rid")) if m_rid else current

        # DRY RUN → emit
        if DRY_RE.search(line):
            ts = _line_ts(line, ctx.get("timestamp_utc"))
            if _in_window(ts):
                yield {
                    "timestamp_utc": ts,
                    "fusion_alias": ctx.get("fusion_alias", ""),
                    "display_name": ctx.get("display_name", ""),
                    "job_definition_name": ctx.get("job_definition_name", ""),
                    "job_definition_id": ctx.get("job_definition_id", ""),
                    "job_package_name": ctx.get("job_package_name", ""),
                    "application": ctx.get("application", ""),
                    "request_id": ctx.get("request_id", ""),
                    "status": "DRY_RUN",
                    "error": ctx.get("error", ""),
                }
            current = {}
            continue

        # FINAL → emit (word or key=value)
        m_final = FINAL_WORD.search(line) or FINAL_KV.search(line)
        if m_final:
            status = (m_final.group(1) if m_final.lastindex else m_final.group("status")).upper()
            ts = _line_ts(line, ctx.get("timestamp_utc"))
            if _in_window(ts):
                yield {
                    "timestamp_utc": ts,
                    "fusion_alias": ctx.get("fusion_alias", ""),
                    "display_name": ctx.get("display_name", ""),
                    "job_definition_name": ctx.get("job_definition_name", ""),
                    "job_definition_id": ctx.get("job_definition_id", ""),
                    "job_package_name": ctx.get("job_package_name", ""),
                    "application": ctx.get("application", ""),
                    "request_id": ctx.get("request_id", ""),
                    "status": status,
                    "error": ctx.get("error", ""),
                }
            current = {}
            continue

def backfill():
    """
    Main entry point - parse logs and write to Excel.
    
    WORKFLOW:
        1. Open or create Excel workbook
        2. For each log file in LOG_PATHS:
           a. Read file line by line
           b. Parse into job records
           c. Convert UTC to local timezone
           d. Append rows to Excel
           e. Report progress every 10 rows
        3. Auto-size columns
        4. Save workbook (or shadow file if locked)
    
    PROGRESS REPORTING:
        Prints status every 10 rows:
        "[backfill] rows written: 30 (this file), 150 (total)"
    
    SHADOW FILE FALLBACK:
        If Excel locked during save, creates shadow file
        Shadow file can be merged later when main file unlocked
    
    ERROR HANDLING:
        File not found: Skip with warning
        Unicode errors: Ignore with errors="ignore"
        PermissionError on save: Use shadow file
    """
    path = _default_xlsx_path()
    wb, ws, real_path = _ensure_wb_sheet(path)

    total = 0
    for p in LOG_PATHS:
        if not Path(p).exists():
            print(f"[backfill] Skip (not found): {p}")
            continue
        print(f"[backfill] Reading: {p}")
        written = 0
        with open(p, "r", encoding="utf-8", errors="ignore") as f:
            for rec in parse_log(f):
                ld, lt = _local_parts(rec["timestamp_utc"], LOCAL_TZ)
                row = [
                    rec["timestamp_utc"], ld, lt,
                    rec.get("fusion_alias", ""),
                    rec.get("display_name", ""),
                    rec.get("job_definition_name", ""),
                    rec.get("job_definition_id", ""),
                    rec.get("job_package_name", ""),
                    rec.get("application", ""),
                    rec.get("request_id", ""),
                    rec.get("status", ""),
                    rec.get("error", ""),
                    "backfilled",
                ]
                ws.append(row)
                written += 1
                total += 1
                if written % PROGRESS_EVERY == 0:
                    print(f"[backfill] rows written: {written} (this file), {total} (total)")
        _autosize(ws)
        try:
            wb.save(real_path)
        except PermissionError:
            # if workbook got locked mid-run, fallback to shadow
            shadow = real_path.with_name(f"{real_path.stem}__shadow{real_path.suffix}")
            wb.save(shadow)
            print(f"[backfill] workbook locked, wrote to {shadow.name}")
        print(f"[backfill] file complete → {written} rows")
    print(f"[backfill] Done. Total rows written: {total}")

if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass
    backfill()
